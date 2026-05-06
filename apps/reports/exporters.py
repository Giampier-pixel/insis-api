import csv
import io

from openpyxl import Workbook

from apps.reports.data import REPORT_TITLES, build_all_reports, build_report


def _fieldnames(rows):
    if not rows:
        return ["empty"]
    names = []
    for row in rows:
        for key in row.keys():
            if key not in names:
                names.append(key)
    return names


def export_csv(report_type, company):
    rows = build_report(report_type, company)
    output = io.StringIO()
    fieldnames = _fieldnames(rows)
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return output.getvalue().encode("utf-8-sig")


def export_excel(report_type, company):
    workbook = Workbook()
    workbook.remove(workbook.active)

    # XLSX exports include all report tabs so HR users can inspect context in
    # one file, while CSV remains a single report.
    for current_type, rows in build_all_reports(company).items():
        worksheet = workbook.create_sheet(REPORT_TITLES[current_type][:31])
        fieldnames = _fieldnames(rows)
        worksheet.append(fieldnames)
        for row in rows:
            worksheet.append([row.get(field) for field in fieldnames])

    metadata = workbook.create_sheet("Export Metadata")
    metadata.append(["selected_report", report_type])
    metadata.append(["company", company.name])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_report(report_type, file_format, company):
    if file_format == "xlsx":
        return export_excel(report_type, company)
    return export_csv(report_type, company)
